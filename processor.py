import json
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

POSITIONS = {
    'SP': '선발투수', 'MP': '중간투수', 'CP': '마무리투수',
    'C': '포수', '1B': '1루수', '2B': '2루수', '3B': '3루수',
    'SS': '유격수', 'OF': '외야수', 'DH': '지명타자'
}

# 색상 정의
COLOR = {
    "nanum_header": "1B3A6B",   # 나눔 - 진한 파랑
    "dream_header": "6B1B1B",   # 드림 - 진한 빨강
    "nanum_1st":    "D6E4F7",   # 나눔 1위 배경
    "dream_1st":    "F7D6D6",   # 드림 1위 배경
    "highlight":    "FFE699",   # 노란 강조 (1위)
    "pos_header":   "2D2D2D",   # 포지션 헤더 배경
    "white":        "FFFFFF",
    "light_gray":   "F5F5F5",
}

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def load_latest_two():
    """가장 최근 2개 JSON 파일 로드 (비교용)"""
    files = sorted(glob.glob("data/*.json"))
    if not files:
        raise FileNotFoundError("data/ 폴더에 JSON 파일이 없습니다.")
    latest = json.load(open(files[-1], encoding="utf-8"))
    prev = json.load(open(files[-2], encoding="utf-8")) if len(files) >= 2 else None
    return latest, prev

def get_vote_map(data):
    """이전 데이터에서 {pos_id: {P_NM: VOTE_CN}} 맵 생성"""
    vote_map = {}

    for pos_id, pos_data in data["positions"].items():
        vote_map[pos_id] = {}

        total_votes = 0

        for team_key in ("nanum", "dream"):
            for p in pos_data.get(team_key, []):

                try:
                    votes = int(p.get("VOTE_CN", "0").replace(",", ""))
                except:
                    votes = 0

                name = p.get("P_NM", "")

                vote_map[pos_id][name] = votes
                total_votes += votes

        vote_map[pos_id]["_total"] = total_votes

    return vote_map

def calc_total_votes(pos_data):
    """포지션 내 전체 득표수 합산"""
    total = 0
    for team_key in ("nanum", "dream"):
        for p in pos_data.get(team_key, []):
            try:
                total += int(p.get("VOTE_CN", "0").replace(",", ""))
            except:
                pass
    return total

def write_header(ws, timestamp):
    ws.merge_cells("A1:T1")
    cell = ws["A1"]
    cell.value = f"2026 신한 SOL KBO 올스타 투표 현황   |   기준: {timestamp[:16].replace('T', ' ')} KST"
    cell.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=COLOR["pos_header"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

def write_position_sheet(ws, pos_name, pos_data, prev_map, pos_id):
    ws.title = pos_name

    nanum_players = pos_data.get("nanum", [])
    dream_players = pos_data.get("dream", [])
    total = calc_total_votes(pos_data)

    # --- 헤더 ---
    ws.merge_cells("A1:G1")
    ws["A1"] = "▶ 나눔 올스타"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR["nanum_header"])
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("I1:O1")
    ws["I1"] = "▶ 드림 올스타"
    ws["I1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    ws["I1"].fill = PatternFill("solid", fgColor=COLOR["dream_header"])
    ws["I1"].alignment = Alignment(horizontal="center")

    col_headers = [
        "순위",
        "선수명",
        "구단",
        "득표수",
        "득표수 증감",
        "득표율",
        "득표율 증감"
    ]

    for i, h in enumerate(col_headers):
        for col_offset, fill_color in [
            (0, COLOR["nanum_header"]),
            (8, COLOR["dream_header"])
        ]:
            c = ws.cell(row=2, column=i + 1 + col_offset, value=h)
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=fill_color)
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border()

    ws.row_dimensions[2].height = 20

    def write_players(players, col_start):
        for idx, p in enumerate(players):

            row = idx + 3

            rank = int(p.get("RANK_CN", 0))
            name = p.get("P_NM", "")
            team = p.get("T_NM", "")

            try:
                votes = int(p.get("VOTE_CN", "0").replace(",", ""))
            except:
                votes = 0

            rate = votes / total * 100 if total > 0 else 0

            prev_votes = prev_map.get(name)
            prev_total = prev_map.get("_total", 0)

            if prev_votes is None or prev_total == 0:
                vote_diff_text = "-"
                rate_diff_text = "-"
            else:
                vote_diff = votes - prev_votes

                prev_rate = prev_votes / prev_total * 100
                rate_diff = rate - prev_rate

                vote_diff_text = f"{vote_diff:+,}"
                rate_diff_text = f"{rate_diff:+.2f}%p"

            is_first = rank == 1

            bg = (
                COLOR["highlight"]
                if is_first
                else ("FFFFFF" if idx % 2 == 0 else COLOR["light_gray"])
            )

            vals = [
                f"{rank}위",
                name,
                team,
                f"{votes:,}",
                vote_diff_text,
                f"{rate:.2f}%",
                rate_diff_text,
            ]

            for i, v in enumerate(vals):
                c = ws.cell(row=row, column=col_start + i, value=v)
                c.font = Font(name="Arial", bold=is_first, size=10)
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="center")
                c.border = thin_border()

            ws.row_dimensions[row].height = 18

    write_players(nanum_players, 1)
    write_players(dream_players, 9)

    # 구분 열
    ws.column_dimensions["H"].width = 2

    # 열 너비
    for col, width in zip(
        "ABCDEFG",
        [6, 10, 8, 12, 12, 9, 12]
    ):
        ws.column_dimensions[col].width = width

    for col, width in zip(
        "IJKLMNO",
        [6, 10, 8, 12, 12, 9, 12]
    ):
        ws.column_dimensions[col].width = width
def build_excel(latest, prev):
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    prev_map = get_vote_map(prev) if prev else {}
    timestamp = latest.get("timestamp", "")
    timestamp_safe = timestamp.replace(":", "-").replace("T", "_") if timestamp else "latest"

    for pos_id, pos_name in POSITIONS.items():
        pos_data = latest["positions"].get(pos_id)
        if not pos_data:
            continue
        ws = wb.create_sheet(title=pos_name)
        write_position_sheet(ws, pos_name, pos_data, prev_map.get(pos_id, {}), pos_id)

    os.makedirs("output", exist_ok=True)
    outpath = f"output/kbo_allstar.xlsx"
    wb.save(outpath)
    print(f"✅ 엑셀 저장 완료: {outpath}")

if __name__ == "__main__":
    latest, prev = load_latest_two()
    build_excel(latest, prev)